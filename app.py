"""
Nokia Config IP Manager — Flask App
별도 프로젝트로 운영하다가 나중에 nokia-config-visualizer와 통합 예정.
"""
import io
import os
import csv
import json
import tempfile
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_file, session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from parser.ip_parser import parse_all_configs

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24).hex())

# 기본 config 디렉토리 (환경변수 CONFIG_DIR 우선, 없으면 실행 위치 기준)
DEFAULT_CONFIG_DIR = os.environ.get('CONFIG_DIR', str(Path(__file__).parent.parent / 'config'))

# 메모리 캐시 (재파싱 방지)
_cache: dict = {'dir': None, 'records': []}


def get_records(config_dir: str) -> list[dict]:
    """파싱 결과 반환 (캐시 활용)"""
    global _cache
    config_dir = str(Path(config_dir).resolve())
    if _cache['dir'] != config_dir:
        _cache['dir'] = config_dir
        _cache['records'] = parse_all_configs(config_dir)
    return _cache['records']


def build_summary(records: list[dict]) -> dict:
    """통계 요약 계산 (by_model은 장비 수 기준)"""
    total = len(records)
    by_type = {}
    by_model = {}
    devices: dict[str, str] = {}  # device_name → model
    latest_date = ''

    for r in records:
        t = r['ip_type']
        by_type[t] = by_type.get(t, 0) + 1

        dn = r['device_name']
        if dn and dn not in devices:
            devices[dn] = r['device_model'] or 'Unknown'

        d = r['config_date']
        if d and d > latest_date:
            latest_date = d

    for model in devices.values():
        by_model[model] = by_model.get(model, 0) + 1

    return {
        'total_ips': total,
        'total_devices': len(devices),
        'by_type': by_type,
        'by_model': by_model,
        'latest_date': latest_date,
    }


# ─────────────────────────────────────────────
# 컬럼 정의 (프론트엔드 key → 헤더명, 값추출, 열너비)
# ─────────────────────────────────────────────
COLUMN_MAP = {
    'network_address': ('네트워크 주소', lambda r: r['network_address'],                                      16),
    'subnet_mask':     ('서브넷 마스크', lambda r: r['subnet_mask'],                                          16),
    'cidr':            ('CIDR',         lambda r: r['cidr'],                                                  18),
    'ip_type':         ('IP 유형',      lambda r: r['ip_type'],                                               13),
    'device_name':     ('장비명',       lambda r: r['device_name'],                                           32),
    'device_model':    ('장비 모델',    lambda r: r['device_model'],                                          18),
    'location':        ('위치',         lambda r: r['location'],                                              20),
    'interface_name':  ('인터페이스',   lambda r: r['interface_name'],                                        14),
    'port':            ('포트',         lambda r: r['port'],                                                  10),
    'desc':            ('설명',         lambda r: r['interface_desc'] or r['port_desc'] or r['route_desc'],   30),
    'peer_device':     ('Peer 장비',    lambda r: r['peer_device'],                                           30),
    'peer_port':       ('Peer 포트',    lambda r: r['peer_port'],                                             10),
    'next_hop_ip':     ('Next-hop IP',  lambda r: r['next_hop_ip'],                                           14),
    'admin_state':     ('상태',         lambda r: r['admin_state'],                                            8),
    'config_date':     ('Config 날짜',  lambda r: r['config_date'],                                           12),
    'router_id':       ('Router ID',    lambda r: r['router_id'],                                             14),
    'as_number':       ('AS 번호',      lambda r: r['as_number'],                                             10),
    'os_version':      ('OS 버전',      lambda r: r['os_version'],                                            16),
    'filename':        ('파일명',       lambda r: r['filename'],                                               40),
}

def resolve_export_cols(cols_param: str) -> list[str]:
    """쿼리 파라미터 cols=key1,key2,... → 유효한 키 목록 반환 (없으면 기본 순서)"""
    if cols_param:
        keys = [k.strip() for k in cols_param.split(',') if k.strip() in COLUMN_MAP]
        if keys:
            return keys
    return list(COLUMN_MAP.keys())


# ─────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html', default_dir=DEFAULT_CONFIG_DIR)


@app.route('/api/load', methods=['POST'])
def api_load():
    """config 디렉토리 지정 후 파싱"""
    data = request.get_json(silent=True) or {}
    config_dir = data.get('config_dir', '').strip()

    if not config_dir:
        config_dir = DEFAULT_CONFIG_DIR

    # Path Traversal 방지
    config_dir = str(Path(config_dir).resolve())
    if not Path(config_dir).is_dir():
        return jsonify({'error': f'디렉토리를 찾을 수 없습니다: {config_dir}'}), 400

    try:
        records = get_records(config_dir)
        summary = build_summary(records)
        return jsonify({
            'success': True,
            'config_dir': config_dir,
            'summary': summary,
            'records': records,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/upload', methods=['POST'])
def api_upload():
    """클라이언트 PC에서 업로드된 config 파일 파싱"""
    global _cache
    files = request.files.getlist('files')
    txt_files = [f for f in files if f.filename.lower().endswith('.txt')]

    if not txt_files:
        return jsonify({'error': '.txt 파일이 없습니다'}), 400

    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            for f in txt_files:
                safe_name = Path(f.filename).name  # 경로 없이 파일명만 사용
                f.save(os.path.join(tmpdir, safe_name))
            records = parse_all_configs(tmpdir)

        # 이후 export 요청을 위해 캐시에 저장
        _cache['dir'] = '__upload__'
        _cache['records'] = records

        summary = build_summary(records)
        return jsonify({
            'success': True,
            'file_count': len(txt_files),
            'summary': summary,
            'records': records,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/browse')
def api_browse():
    """서버 파일시스템 디렉토리 탐색"""
    req_path = request.args.get('path', '').strip()

    # 시작 경로: 요청 경로 → DEFAULT_CONFIG_DIR → 홈 디렉토리
    if req_path:
        base = Path(req_path).resolve()
    else:
        base = Path(DEFAULT_CONFIG_DIR).resolve()
        if not base.is_dir():
            base = Path.home()

    if not base.is_dir():
        base = Path.home()

    # 상위 디렉토리 (루트면 None)
    parent = str(base.parent) if base.parent != base else None

    # 하위 디렉토리 목록 (숨김 제외, 이름순 정렬)
    dirs = []
    try:
        for d in sorted(base.iterdir(), key=lambda x: x.name.lower()):
            if d.is_dir() and not d.name.startswith('.'):
                dirs.append({'name': d.name, 'path': str(d)})
    except PermissionError:
        pass

    return jsonify({
        'current': str(base),
        'parts':   [{'name': p or '/', 'path': str(Path(*base.parts[:i+1]))}
                    for i, p in enumerate(base.parts)],
        'parent':  parent,
        'dirs':    dirs,
    })


@app.route('/api/data')
def api_data():
    """현재 로드된 데이터 반환"""
    config_dir = request.args.get('dir', DEFAULT_CONFIG_DIR).strip()
    config_dir = str(Path(config_dir).resolve())

    if not Path(config_dir).is_dir():
        return jsonify({'error': '디렉토리 없음'}), 400

    records = get_records(config_dir)
    summary = build_summary(records)
    return jsonify({'summary': summary, 'records': records})


@app.route('/api/export/excel')
def export_excel():
    """Excel 다운로드 (시트 분리: 전체, Interface IP, Static Route, 장비목록)"""
    config_dir = request.args.get('dir', DEFAULT_CONFIG_DIR).strip()
    if config_dir == '__upload__':
        records = _cache['records']
        if not records:
            return jsonify({'error': '먼저 파일을 업로드하세요'}), 400
    else:
        config_dir = str(Path(config_dir).resolve())
        records = get_records(config_dir)

    wb = openpyxl.Workbook()

    # 공통 스타일
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    header_fill_blue   = PatternFill('solid', fgColor='1A6EBD')
    header_fill_green  = PatternFill('solid', fgColor='1A7A4A')
    header_fill_orange = PatternFill('solid', fgColor='C25B00')
    header_fill_gray   = PatternFill('solid', fgColor='555555')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin   = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_sheet(ws, rows, headers, col_widths, header_fill):
        # 헤더
        ws.append(headers)
        for col_idx, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[1].height = 30

        # 데이터
        for row in rows:
            ws.append(row)

        # 행 스타일
        for row_idx in range(2, ws.max_row + 1):
            fill = PatternFill('solid', fgColor='F8F9FA' if row_idx % 2 == 0 else 'FFFFFF')
            for cell in ws[row_idx]:
                cell.fill = fill
                cell.alignment = left
                cell.border = border
            ws.row_dimensions[row_idx].height = 18

        # 열 너비
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 틀 고정
        ws.freeze_panes = 'A2'

    # ── Sheet 1: 사용자 설정 컬럼 순서/표시 ──
    col_keys = resolve_export_cols(request.args.get('cols', ''))
    ws1 = wb.active
    ws1.title = '전체 IP 목록'
    headers_user  = [COLUMN_MAP[k][0] for k in col_keys]
    col_widths_user = [COLUMN_MAP[k][2] for k in col_keys]
    rows_user = [
        [COLUMN_MAP[k][1](r) for k in col_keys]
        for r in records
    ]
    write_sheet(ws1, rows_user, headers_user, col_widths_user, header_fill_blue)

    # ── Sheet 2: Interface IP ──
    ws2 = wb.create_sheet('Interface IP')
    iface_records = [r for r in records if r['ip_type'] == 'Interface IP']
    headers_iface = [
        'CIDR', 'IP 주소', 'Prefix', '서브넷 마스크', '네트워크 주소',
        '장비명', '장비 모델', '위치', '인터페이스', '포트',
        'Peer 장비', 'Peer 포트', '설명', '상태', 'Config 날짜'
    ]
    rows_iface = [
        [r['cidr'], r['ip_address'], f"/{r['prefix_length']}", r['subnet_mask'],
         r['network_address'], r['device_name'], r['device_model'], r['location'],
         r['interface_name'], r['port'], r['peer_device'], r['peer_port'],
         r['interface_desc'] or r['port_desc'], r['admin_state'], r['config_date']]
        for r in iface_records
    ]
    write_sheet(ws2, rows_iface, headers_iface,
                [18,14,7,16,16,32,18,20,12,10,30,10,35,8,12],
                header_fill_green)

    # ── Sheet 3: Static Route ──
    ws3 = wb.create_sheet('Static Route')
    sr_records = [r for r in records if r['ip_type'] == 'Static Route']
    headers_sr = [
        '목적지 CIDR', '네트워크 주소', 'Prefix', '서브넷 마스크',
        '장비명', '장비 모델', '위치',
        'Next-hop IP', 'Peer 장비', 'Peer 포트', '설명', '상태', 'Config 날짜'
    ]
    rows_sr = [
        [r['cidr'], r['network_address'], f"/{r['prefix_length']}", r['subnet_mask'],
         r['device_name'], r['device_model'], r['location'],
         r['next_hop_ip'], r['peer_device'], r['peer_port'],
         r['route_desc'], r['admin_state'], r['config_date']]
        for r in sr_records
    ]
    write_sheet(ws3, rows_sr, headers_sr,
                [18,16,7,16,32,18,20,14,30,10,35,8,12],
                header_fill_orange)

    # ── Sheet 4: 장비 목록 ──
    ws4 = wb.create_sheet('장비 목록')
    devices_seen = {}
    for r in records:
        dn = r['device_name']
        if dn not in devices_seen:
            devices_seen[dn] = r
    headers_dev = [
        '장비명', '장비 모델', '위치', 'System IP', 'Router ID', 'AS 번호', 'OS 버전', 'Config 날짜', '파일명'
    ]
    rows_dev = []
    for dn, r in sorted(devices_seen.items()):
        sys_ip = next(
            (x['cidr'] for x in records if x['device_name'] == dn and x['ip_type'] == 'System IP'),
            ''
        )
        rows_dev.append([
            r['device_name'], r['device_model'], r['location'],
            sys_ip, r['router_id'], r['as_number'],
            r['os_version'], r['config_date'], r['filename']
        ])
    write_sheet(ws4, rows_dev, headers_dev,
                [32,18,22,16,14,10,18,12,42],
                header_fill_gray)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='IP_관리대장.xlsx'
    )


@app.route('/api/export/csv')
def export_csv():
    """CSV 다운로드 (UTF-8 BOM)"""
    config_dir = request.args.get('dir', DEFAULT_CONFIG_DIR).strip()
    if config_dir == '__upload__':
        records = _cache['records']
        if not records:
            return jsonify({'error': '먼저 파일을 업로드하세요'}), 400
    else:
        config_dir = str(Path(config_dir).resolve())
        records = get_records(config_dir)

    col_keys = resolve_export_cols(request.args.get('cols', ''))

    output = io.StringIO()
    output.write('\ufeff')  # UTF-8 BOM (Excel 한글 호환)
    writer = csv.writer(output)
    writer.writerow([COLUMN_MAP[k][0] for k in col_keys])
    for r in records:
        writer.writerow([COLUMN_MAP[k][1](r) for k in col_keys])

    buf = io.BytesIO(output.getvalue().encode('utf-8-sig'))
    return send_file(
        buf,
        mimetype='text/csv; charset=utf-8',
        as_attachment=True,
        download_name='IP_관리대장.csv'
    )


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)
